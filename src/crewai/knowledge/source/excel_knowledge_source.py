from pathlib import Path
from typing import Dict, List

from crewai.knowledge.source.base_file_knowledge_source import BaseFileKnowledgeSource


class ExcelKnowledgeSource(BaseFileKnowledgeSource):
    """A knowledge source that stores and queries Excel file content using embeddings."""

    def load_content(self) -> Dict[Path, str]:
        """Load and preprocess Excel file content. Updated to account for .xlsx workbooks with multiple tabs/sheets"""
        pd, openpyxl, load_workbook = self._import_dependencies()

        # Initialize the content dictionary
        content_dict = {}
        for file_path in self.safe_file_paths:
            # Convert the file path to a Path object
            file_path = self.convert_to_path(file_path)
            # Load the Excel file
            wb = load_workbook(file_path)
            # Get the sheet names
            sheet_names = wb.sheetnames
            # Iterate over the sheets
            # Initialize the file sheet dictionary
            sheet_dict = {}
            for sheet_name in sheet_names:
                # Get the sheet
                ws = wb[sheet_name]
                # Convert the sheet to a CSV string
                sheet_str = """"""
                for row in ws.values:
                    for cell in row:
                        sheet_str += str(cell) + ","
                    sheet_str += "\n"

                print(sheet_str)
                # Add the sheet content to the file sheet dictionary
                sheet_dict[sheet_name] = sheet_str
            # Add the file sheet dictionary to the content dictionary
            content_dict[file_path] = sheet_dict

        return content_dict

    def _import_dependencies(self):
        """Dynamically import dependencies."""
        try:
            import openpyxl  # noqa
            from openpyxl import load_workbook
            import pandas as pd

            return pd, openpyxl, load_workbook
        except ImportError as e:
            missing_package = str(e).split()[-1]
            raise ImportError(
                f"{missing_package} is not installed. Please install it with: pip install {missing_package}"
            )

    def add(self) -> None:
        """
        Add Excel file content to the knowledge source, chunk it, compute embeddings,
        and save the embeddings.
        """
        # Convert dictionary values to a single string if content is a dictionary
        # Updated to account for .xlsx workbooks with multiple tabs/sheets
        content_str = ""
        for value in self.content.values():
            if isinstance(value, dict):
                for sheet_value in value.values():
                    content_str += str(sheet_value) + "\n"
            else:
                content_str += str(value) + "\n"

        new_chunks = self._chunk_text(content_str)
        self.chunks.extend(new_chunks)
        self._save_documents()

    def _chunk_text(self, text: str) -> List[str]:
        """Utility method to split text into chunks."""
        return [
            text[i : i + self.chunk_size]
            for i in range(0, len(text), self.chunk_size - self.chunk_overlap)
        ]
