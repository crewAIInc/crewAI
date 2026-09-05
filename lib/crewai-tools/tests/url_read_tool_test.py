import time
from io import BytesIO
from unittest.mock import patch
import zipfile

import pytest
import re
import requests

from crewai_tools import URLReadTool
from crewai_tools.security.safe_requests import safe_get_bounded


TOOL_MODULE = "crewai_tools.tools.url_read_tool.url_read_tool"


class FakeResponse:
    """Minimal stand-in for a streamed requests.Response."""

    def __init__(
        self,
        body: bytes = b"",
        content_type: str = "text/plain",
        url: str = "https://example.com/file.txt",
        status_code: int = 200,
        chunk_size: int | None = None,
    ):
        self._body = body
        self._chunk_size = chunk_size
        self.headers = {"Content-Type": content_type} if content_type else {}
        self.url = url
        self.status_code = status_code
        self.history: list["FakeResponse"] = []
        self.closed = False

    def raise_for_status(self) -> None:
        """Mimic requests' error-status behavior."""
        if self.status_code >= 400:
            raise requests.HTTPError(f"{self.status_code} error")

    def iter_content(self, chunk_size: int = 65536):
        """Yield the body in chunks, like a streamed response."""
        size = self._chunk_size or chunk_size
        for index in range(0, len(self._body), size):
            yield self._body[index : index + size]

    def close(self) -> None:
        """Record that the response was closed."""
        self.closed = True


def build_pdf(text: str = "Quarterly revenue was 42") -> bytes:
    """Return the bytes of a one-page PDF containing *text*."""
    pymupdf = pytest.importorskip("pymupdf")
    document = pymupdf.open()
    document.new_page().insert_text((72, 72), text)
    try:
        return document.tobytes()
    finally:
        document.close()


PRESIGNED_URL = (
    "https://temp.4d4f16c61d89ec64e760039c4ec50717.r2.cloudflarestorage.com/"
    "668641/share_point/SHARE_POINT_DOWNLOAD_FILE_BY_SERVER_RELATIVE_URL/"
    "response/34e077085d293bdb832a6b7c93b9e222"
    "?X-Amz-Algorithm=AWS4-HMAC-SHA256&X-Amz-Signature=1954b3d4"
)


def build_docx(text: str = "Signed and delivered") -> bytes:
    """Return the bytes of a DOCX holding a single paragraph."""
    from docx import Document

    document = Document()
    document.add_paragraph(text)
    buffer = BytesIO()
    document.save(buffer)
    return buffer.getvalue()


def build_xlsx(rows: list[list[object]], title: str = "Sheet1") -> bytes:
    """Return the bytes of a single-sheet XLSX holding *rows*."""
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = title
    for row in rows:
        worksheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_forged_dimension_xlsx() -> bytes:
    """Return a tiny XLSX whose sheet declares Excel's maximum dimension.

    openpyxl trusts the declared width and pads every row out to it, so this
    4.8 KB file otherwise drives ~1.6e9 cell normalizations.
    """
    from openpyxl import Workbook

    source = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "header"
    worksheet["B100000"] = "stray"
    workbook.save(source)
    workbook.close()

    rewritten = BytesIO()
    with (
        zipfile.ZipFile(BytesIO(source.getvalue())) as archive,
        zipfile.ZipFile(rewritten, "w", zipfile.ZIP_DEFLATED) as output,
    ):
        for info in archive.infolist():
            payload = archive.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                payload = re.sub(
                    rb'<dimension ref="[^"]*"',
                    b'<dimension ref="A1:XFD1048576"',
                    payload,
                )
            output.writestr(info, payload)
    return rewritten.getvalue()


def build_zip(*names: str) -> bytes:
    """Return a zip holding *names*, shaped like an OOXML package."""
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w") as archive:
        for name in names:
            archive.writestr(name, "<x/>")
    return buffer.getvalue()


def fetch_result(
    body: bytes,
    content_type: str = "text/plain",
    url: str = "https://example.com/f.txt",
):
    """Build the (body, content_type, final_url) tuple safe_get_bounded returns."""
    return body, content_type, url


def test_reads_plain_text():
    """A text response is returned as-is, with the configured limits applied."""
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(b"hello world")
        assert tool.run(url="https://example.com/f.txt") == "hello world"

    assert fetch.call_args.kwargs["max_bytes"] == 5 * 1024 * 1024
    assert fetch.call_args.kwargs["timeout"] == 30


def test_honors_declared_charset():
    """The charset in the Content-Type header drives decoding."""
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(
            "café".encode("latin-1"), "text/plain; charset=iso-8859-1"
        )
        assert tool.run(url="https://example.com/f.txt") == "café"


def test_encoding_override_wins_over_server_charset():
    """An explicit encoding beats whatever the server declares."""
    tool = URLReadTool(encoding="latin-1")
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(
            "café".encode("latin-1"), "text/plain; charset=utf-8"
        )
        assert tool.run(url="https://example.com/f.txt") == "café"


def test_undecodable_bytes_fall_back_instead_of_failing():
    """Partially readable text beats an error for the agent."""
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(b"\xff\xfe bad bytes", "text/plain")
        result = tool.run(url="https://example.com/f.txt")

    assert "bad bytes" in result
    assert not result.startswith("Error:")


def test_line_window():
    """start_line and line_count select a window of the extracted text."""
    tool = URLReadTool()
    body = b"one\ntwo\nthree\nfour\nfive\n"
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(body)
        result = tool.run(url="https://example.com/f.txt", start_line=2, line_count=2)

    assert result == "two\nthree\n"


def test_start_line_past_end_reports_error():
    """Asking past the end of the content is reported, not silently empty."""
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(b"one\ntwo\n")
        result = tool.run(url="https://example.com/f.txt", start_line=99)

    assert "exceeds the number of lines" in result


@pytest.mark.parametrize(
    "line_args",
    [{"line_count": -5}, {"line_count": 0}, {"start_line": 0}, {"start_line": -5}],
)
def test_line_arguments_below_one_are_refused(line_args):
    """Out-of-range line arguments are rejected before any request is made.

    islice raises on a negative stop index, and the windowing runs outside the
    tool's error handling, so these have to be refused at validation time.
    """
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        with pytest.raises(ValueError, match="greater than or equal to 1"):
            tool.run(url="https://example.com/f.txt", **line_args)

    fetch.assert_not_called()


def test_json_is_returned_verbatim():
    """JSON is passed through undecorated so callers can parse it."""
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(b'{"a": 1}', "application/json")
        assert tool.run(url="https://example.com/data.json") == '{"a": 1}'


def test_structured_suffix_type_is_treated_as_text():
    """A +json vendor type is text, not an unsupported binary type."""
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(b'{"a": 1}', "application/vnd.api+json")
        assert tool.run(url="https://example.com/data") == '{"a": 1}'


def test_html_is_stripped_to_visible_text():
    """HTML returns visible text with script and style content removed."""
    tool = URLReadTool()
    body = b"<html><head><style>p{color:red}</style></head><body><p>Hi</p><script>x=1</script></body></html>"
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(body, "text/html; charset=utf-8")
        result = tool.run(url="https://example.com/page")

    assert "Hi" in result
    assert "x=1" not in result
    assert "color:red" not in result


def test_binary_content_type_is_rejected():
    """An unsupported type is refused rather than returned as base64."""
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(b"\x89PNG\r\n", "image/png")
        result = tool.run(url="https://example.com/logo.png")

    assert "Unsupported content type 'image/png'" in result


def test_octet_stream_pdf_falls_back_to_url_extension():
    """A PDF served as octet-stream is still extracted, via its extension."""
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(
            build_pdf("Fallback worked"),
            "application/octet-stream",
            "https://example.com/a/b.pdf",
        )
        result = tool.run(url="https://example.com/a/b.pdf")

    assert "Fallback worked" in result


def test_missing_content_type_falls_back_to_url_extension():
    """No Content-Type at all still reads as text when the path says .csv."""
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(b"a,b\n1,2\n", "", "https://example.com/b.csv")
        assert tool.run(url="https://example.com/b.csv") == "a,b\n1,2\n"


def test_query_string_does_not_break_extension_fallback():
    """A presigned-style query string does not hide the path's extension."""
    tool = URLReadTool()
    url = "https://example.com/b.csv?X-Amz-Signature=abc"
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(b"a,b\n", "application/octet-stream", url)
        assert tool.run(url=url) == "a,b\n"


def test_extension_from_requested_url_survives_a_redirect():
    """A .pdf link that redirects to an extensionless path is still extracted.

    Presigned CDN targets routinely drop the extension and serve octet-stream,
    so the requested URL is the only place the type survives.
    """
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(
            build_pdf("Survived the redirect"),
            "application/octet-stream",
            "https://cdn.example.com/objects/9f8a7b6c5d",
        )
        result = tool.run(url="https://example.com/report.pdf")

    assert "Survived the redirect" in result


def test_octet_stream_with_unknown_extension_is_rejected():
    """With neither a usable type nor a known extension, the read is refused."""
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(
            b"\x00\x01", "application/octet-stream", "https://example.com/a/b.bin"
        )
        result = tool.run(url="https://example.com/a/b.bin")

    assert "Unsupported content type" in result


def test_validation_failure_is_returned_as_error():
    """An SSRF rejection reaches the agent as an error string, not an exception."""
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.side_effect = ValueError(
            "URL 'http://169.254.169.254/' resolves to private/reserved IP 169.254.169.254."
        )
        result = tool.run(url="http://169.254.169.254/")

    assert result.startswith("Error:")
    assert "private/reserved IP" in result


def test_request_failure_is_returned_as_error():
    """A transport failure is reported without raising out of the tool."""
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.side_effect = requests.ConnectionError("connection refused")
        result = tool.run(url="https://example.com/f.txt")

    assert result.startswith("Error: Failed to fetch")


def test_custom_headers_are_merged_over_defaults():
    """Caller headers win, but the default User-Agent survives."""
    tool = URLReadTool(headers={"Authorization": "Bearer x", "Accept": "text/plain"})
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(b"ok")
        tool.run(url="https://example.com/f.txt")

    headers = fetch.call_args.kwargs["headers"]
    assert headers["Authorization"] == "Bearer x"
    assert headers["Accept"] == "text/plain"
    assert "crewai-tools URLReadTool" in headers["User-Agent"]


def test_reads_a_real_pdf_end_to_end():
    """Real PDF bytes are extracted page by page."""
    pdf_bytes = build_pdf()

    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(
            pdf_bytes, "application/pdf", "https://example.com/report.pdf"
        )
        result = tool.run(url="https://example.com/report.pdf")

    assert "Page 1:" in result
    assert "Quarterly revenue was 42" in result


def test_corrupt_pdf_reports_error_without_raising():
    """A malformed PDF becomes an error string, not a traceback."""
    pytest.importorskip("pymupdf")

    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(
            b"%PDF-1.4 not really a pdf", "application/pdf"
        )
        result = tool.run(url="https://example.com/report.pdf")

    assert result.startswith("Error: Failed to read PDF content")


def test_presigned_octet_stream_pdf_is_read_from_its_bytes():
    """The reported failure: octet-stream, no extension, real PDF bytes.

    Presigned object-store links from the SharePoint connector use a content
    hash for a path and pin every object to octet-stream, which leaves the
    body as the only evidence of what was fetched.
    """
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(
            build_pdf("Signed quarterly report"),
            "application/octet-stream",
            PRESIGNED_URL,
        )
        result = tool.run(url=PRESIGNED_URL)

    assert "Page 1:" in result
    assert "Signed quarterly report" in result


def test_presigned_octet_stream_docx_is_read_from_its_bytes():
    """A DOCX behind the same extensionless presigned link is extracted."""
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(
            build_docx("Countersigned on Tuesday"),
            "application/octet-stream",
            PRESIGNED_URL,
        )
        result = tool.run(url=PRESIGNED_URL)

    assert result == "Countersigned on Tuesday"


def test_octet_stream_html_is_sniffed_and_stripped():
    """HTML bytes behind an unhelpful header still lose their markup."""
    tool = URLReadTool()
    body = b"<!DOCTYPE html><html><body><p>Hi</p><script>x=1</script></body></html>"
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(body, "application/octet-stream", PRESIGNED_URL)
        result = tool.run(url=PRESIGNED_URL)

    assert "Hi" in result
    assert "x=1" not in result


@pytest.mark.parametrize(
    "prefix",
    [
        pytest.param(b"", id="bare"),
        pytest.param(b"\xef\xbb\xbf", id="utf8-bom"),
        pytest.param(b"\n  \t", id="leading-whitespace"),
        pytest.param(b"\xef\xbb\xbf\n  ", id="bom-then-whitespace"),
    ],
)
def test_bom_and_whitespace_do_not_hide_the_html_prefix(prefix):
    """A BOM or leading whitespace must not demote HTML to raw text."""
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(
            prefix + b"<html><body><p>Hi</p></body></html>",
            "application/octet-stream",
            PRESIGNED_URL,
        )
        result = tool.run(url=PRESIGNED_URL)

    assert "<p>" not in result
    assert "Hi" in result


@pytest.mark.parametrize(
    "body",
    [
        pytest.param(b"a,b\n1,2\n", id="csv"),
        pytest.param(b'{"a": 1}', id="json"),
        pytest.param(b"# Title\n\nBody text.\n", id="markdown"),
        pytest.param("plain café text\n".encode(), id="utf8-plain"),
    ],
)
def test_octet_stream_text_bodies_are_returned_verbatim(body):
    """Decodable, NUL-free bytes are handed back untouched."""
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(body, "application/octet-stream", PRESIGNED_URL)
        assert tool.run(url=PRESIGNED_URL) == body.decode()


@pytest.mark.parametrize(
    "entries",
    [
        pytest.param(("[Content_Types].xml", "ppt/presentation.xml"), id="pptx"),
        pytest.param(("[Content_Types].xml", "visio/document.xml"), id="vsdx"),
        pytest.param(("notes.txt",), id="plain-zip"),
    ],
)
def test_unsupported_zips_are_refused_without_a_misleading_docx_error(entries):
    """A .pptx must get an honest refusal, not a DOCX extraction failure.

    Sniffing the zip magic alone would route any OOXML package into
    python-docx, which raises and surfaces as "Failed to read DOCX content"
    -- a worse answer than the refusal it replaced.
    """
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(
            build_zip(*entries), "application/octet-stream", PRESIGNED_URL
        )
        result = tool.run(url=PRESIGNED_URL)

    assert "Unsupported content type 'application/octet-stream'" in result
    assert "Failed to read DOCX" not in result


@pytest.mark.parametrize(
    "body",
    [
        pytest.param(build_docx()[:120], id="truncated-docx"),
        pytest.param(b"PK\x03\x04", id="magic-only"),
        pytest.param(b"PK\x03\x04" + b"\xff" * 200, id="garbage-after-magic"),
    ],
)
def test_malformed_zip_bodies_are_refused_without_raising(body):
    """Zip magic on unreadable bytes fails closed rather than escaping _run."""
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(body, "application/octet-stream", PRESIGNED_URL)
        result = tool.run(url=PRESIGNED_URL)

    assert "Unsupported content type" in result


def test_multibyte_character_past_a_prefix_boundary_still_reads_as_text():
    """The sniff decodes the whole body, so no character is split in half.

    Decoding only a leading slice rejects valid UTF-8 whenever a multi-byte
    character straddles the cut.
    """
    tool = URLReadTool()
    body = b"a" * 2047 + "é".encode() + b"b" * 5000
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(body, "application/octet-stream", PRESIGNED_URL)
        result = tool.run(url=PRESIGNED_URL)

    assert not result.startswith("Error:")
    assert result == body.decode()


def test_empty_body_is_refused_rather_than_read_as_empty_text():
    """An empty body identifies nothing; it must not read as a successful "".

    Without an explicit guard it strict-decodes to "" with no NUL byte and
    would be classified as text.
    """
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(b"", "application/octet-stream", PRESIGNED_URL)
        result = tool.run(url=PRESIGNED_URL)

    assert "Unsupported content type" in result


@pytest.mark.parametrize(
    "body",
    [
        pytest.param("café,x\n".encode("latin-1"), id="latin-1"),
        pytest.param("a,b\n".encode("utf-16"), id="utf-16-with-bom"),
        pytest.param("a,b\n".encode("utf-16-be"), id="utf-16-be-nul-bytes"),
        pytest.param(b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR", id="png"),
    ],
)
def test_undecodable_or_nul_bearing_bodies_fail_closed(body):
    """Fail-closed is deliberate: only strict UTF-8 without NUL reads as text.

    A charset-guessing rescue here would push binary payloads into an agent's
    context as mojibake, which is what the tool's text-only contract forbids.
    """
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(body, "application/octet-stream", PRESIGNED_URL)
        result = tool.run(url=PRESIGNED_URL)

    assert "Unsupported content type 'application/octet-stream'" in result


def test_declared_content_type_wins_over_the_body_bytes():
    """A usable header is still authoritative; the sniff never overrides it."""
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(
            build_pdf("Should not be extracted"), "text/html", PRESIGNED_URL
        )
        result = tool.run(url=PRESIGNED_URL)

    assert "Page 1:" not in result
    assert "Should not be extracted" not in result


def test_url_extension_wins_over_the_body_bytes():
    """The extension fallback still runs ahead of the sniff."""
    tool = URLReadTool()
    url = "https://example.com/export.csv"
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(
            build_pdf("Should not be extracted"), "application/octet-stream", url
        )
        result = tool.run(url=url)

    assert result.startswith("%PDF")
    assert "Page 1:" not in result


def test_sniffed_content_still_honors_the_line_window():
    """Windowing applies to sniffed bodies like any other."""
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(
            b"one\ntwo\nthree\nfour\n", "application/octet-stream", PRESIGNED_URL
        )
        result = tool.run(url=PRESIGNED_URL, start_line=2, line_count=2)

    assert result == "two\nthree\n"


def test_presigned_octet_stream_xlsx_is_read_from_its_bytes():
    """The reported file: an XLSX behind an extensionless presigned link."""
    tool = URLReadTool()
    body = build_xlsx([["RFQ ID", "Title"], ["RFQ-1", "Turbine parts"]], "RFQ Header")
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(body, "application/octet-stream", PRESIGNED_URL)
        result = tool.run(url=PRESIGNED_URL)

    assert result == "Sheet RFQ Header:\nRFQ ID,Title\nRFQ-1,Turbine parts"


@pytest.mark.parametrize(
    ("content_type", "url"),
    [
        pytest.param(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            PRESIGNED_URL,
            id="declared-type",
        ),
        pytest.param(
            "application/octet-stream",
            "https://example.com/q3.xlsx",
            id="url-extension",
        ),
    ],
)
def test_xlsx_resolves_from_its_declared_type_and_its_extension(content_type, url):
    """XLSX is reachable by all three routes, not only by sniffing."""
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(build_xlsx([["a", "b"]]), content_type, url)
        assert tool.run(url=url) == "Sheet Sheet1:\na,b"


def test_xlsx_cells_are_csv_quoted_so_the_grid_survives():
    """A comma, quote or newline inside a cell must not corrupt the row."""
    tool = URLReadTool()
    body = build_xlsx([["Smith, Jane", 'He said "hi"'], ["line1\nline2", "plain"]])
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(body, "application/octet-stream", PRESIGNED_URL)
        result = tool.run(url=PRESIGNED_URL)

    assert '"Smith, Jane"' in result
    assert '"He said ""hi"""' in result
    assert '"line1\nline2"' in result


def test_xlsx_blank_rows_are_dropped():
    """Excel reports generous dimensions; phantom rows must not pad the output."""
    tool = URLReadTool()
    body = build_xlsx([["a"], [None], ["b"], [None], [None]])
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(body, "application/octet-stream", PRESIGNED_URL)
        result = tool.run(url=PRESIGNED_URL)

    assert result == "Sheet Sheet1:\na\nb"


def test_one_far_down_cell_does_not_pad_the_output():
    """A stray cell at row 100000 must not expand 5 KB into 100k blank rows.

    openpyxl pads every row up to the sheet's declared dimension, so trimming
    only trailing blanks left the interior padding in the agent's context.
    """
    tool = URLReadTool()
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "header"
    worksheet["B100000"] = "stray"
    buffer = BytesIO()
    workbook.save(buffer)
    body = buffer.getvalue()

    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(body, "application/octet-stream", PRESIGNED_URL)
        result = tool.run(url=PRESIGNED_URL)

    assert len(result.splitlines()) == 3
    assert "header" in result
    assert "stray" in result


def test_oversized_workbook_is_truncated_with_a_visible_notice():
    """A cap that is not announced reads as complete content. Announce it."""
    tool = URLReadTool()
    body = build_xlsx([[f"r{index}c{column}" for column in range(10)] for index in range(30)])
    with (
        patch(f"{TOOL_MODULE}._XLSX_MAX_CELLS", 50),
        patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch,
    ):
        fetch.return_value = fetch_result(body, "application/octet-stream", PRESIGNED_URL)
        result = tool.run(url=PRESIGNED_URL)

    assert "[Truncated: workbook is too large to read in full]" in result
    assert "r0c0" in result
    assert "r29c9" not in result


def test_xlsx_whitespace_only_values_survive():
    """Padding is empty, not blank -- a cell the author filled with spaces stays.

    A bare rstrip() on the rendered grid would also eat a trailing space from
    the final cell, and dropping rows on .strip() would delete a row whose
    cells hold only spaces.
    """
    tool = URLReadTool()
    body = build_xlsx([["a", "trailing "], [" ", " "], ["b", "c"]])
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(body, "application/octet-stream", PRESIGNED_URL)
        result = tool.run(url=PRESIGNED_URL)

    assert result == "Sheet Sheet1:\na,trailing \n , \nb,c"


def test_xlsx_trailing_space_in_the_final_cell_survives():
    """The rendered grid loses its line terminator, not the last cell's space."""
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(
            build_xlsx([["only "]]), "application/octet-stream", PRESIGNED_URL
        )
        assert tool.run(url=PRESIGNED_URL) == "Sheet Sheet1:\nonly "


def test_forged_sheet_dimension_is_bounded_by_the_scan_budget():
    """A forged dimension must not buy unbounded work off a 5 KB upload.

    Blank rows are skipped, so budgeting only emitted cells left the padding
    free: 4.8 KB drove 1.6e9 normalizations in 15s. The scan budget is
    charged per row before the row is normalized, which is what bounds it.
    """
    tool = URLReadTool()
    body = build_forged_dimension_xlsx()
    assert len(body) < 10_000

    started = time.monotonic()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(body, "application/octet-stream", PRESIGNED_URL)
        result = tool.run(url=PRESIGNED_URL)
    elapsed = time.monotonic() - started

    assert "[Truncated: workbook is too large to read in full]" in result
    # Generous vs. the ~15s the unbounded scan took, tight enough to fail if
    # the per-row charge is removed.
    assert elapsed < 5, f"scan took {elapsed:.1f}s -- the budget is not bounding work"


def test_zip_claiming_to_be_both_docx_and_xlsx_is_refused():
    """A package asserting two identities has not been positively identified."""
    tool = URLReadTool()
    body = build_zip("[Content_Types].xml", "word/document.xml", "xl/workbook.xml")
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(body, "application/octet-stream", PRESIGNED_URL)
        result = tool.run(url=PRESIGNED_URL)

    assert "Unsupported content type 'application/octet-stream'" in result
    assert "Failed to read" not in result


def test_xlsx_with_no_cells_says_so_instead_of_returning_nothing():
    """An empty workbook reports its emptiness rather than an empty string."""
    tool = URLReadTool()
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(
            build_xlsx([]), "application/octet-stream", PRESIGNED_URL
        )
        assert tool.run(url=PRESIGNED_URL) == "[XLSX with no extractable cells]"


def test_xlsx_formula_without_a_cached_value_reads_as_empty():
    """data_only returns cached results, so an uncalculated formula is blank.

    Pinning this documents the trade: agents get "42" from a workbook Excel
    has saved, never the literal "=SUM(A1:A2)".
    """
    tool = URLReadTool()
    body = build_xlsx([[1], [2], ["=SUM(A1:A2)"]])
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(body, "application/octet-stream", PRESIGNED_URL)
        result = tool.run(url=PRESIGNED_URL)

    assert "=SUM" not in result
    assert result == "Sheet Sheet1:\n1\n2"


def test_corrupt_xlsx_reports_error_without_raising():
    """A zip that claims to be a workbook but is not becomes an error string."""
    tool = URLReadTool()
    body = build_zip("[Content_Types].xml", "xl/workbook.xml")
    with patch(f"{TOOL_MODULE}.safe_get_bounded") as fetch:
        fetch.return_value = fetch_result(body, "application/octet-stream", PRESIGNED_URL)
        result = tool.run(url=PRESIGNED_URL)

    assert result.startswith("Error: Failed to read XLSX content")


class TestSafeGetBounded:
    """Tests for the bounded-fetch helper itself."""

    def test_returns_body_content_type_and_final_url(self):
        """The helper reports the body alongside where it ended up."""
        response = FakeResponse(b"payload", "text/plain", "https://example.com/final")
        with patch(
            "crewai_tools.security.safe_requests.safe_get", return_value=response
        ):
            body, content_type, final_url = safe_get_bounded(
                "https://example.com/start", max_bytes=1024
            )

        assert body == b"payload"
        assert content_type == "text/plain"
        assert final_url == "https://example.com/final"
        assert response.closed

    def test_rejects_body_over_the_limit(self):
        """Crossing max_bytes raises rather than truncating silently."""
        response = FakeResponse(b"x" * 100, chunk_size=10)
        with patch(
            "crewai_tools.security.safe_requests.safe_get", return_value=response
        ):
            with pytest.raises(ValueError, match="exceeds the 25 byte limit"):
                safe_get_bounded("https://example.com/big", max_bytes=25)

        assert response.closed

    def test_oversized_error_names_the_url_that_served_the_body(self):
        """After a redirect the requested URL is not the one that sent it."""
        response = FakeResponse(
            b"x" * 100, url="https://cdn.example.com/final", chunk_size=10
        )
        with patch(
            "crewai_tools.security.safe_requests.safe_get", return_value=response
        ):
            with pytest.raises(ValueError, match="https://cdn.example.com/final"):
                safe_get_bounded("https://example.com/start", max_bytes=25)

    @pytest.mark.parametrize("max_bytes", [0, -1])
    def test_non_positive_max_bytes_fails_before_requesting(self, max_bytes):
        """A misconfigured cap is caught without issuing a request."""
        with patch("crewai_tools.security.safe_requests.safe_get") as safe_get:
            with pytest.raises(ValueError, match="max_bytes must be positive"):
                safe_get_bounded("https://example.com/f", max_bytes=max_bytes)

        safe_get.assert_not_called()

    def test_stops_reading_once_the_limit_is_crossed(self):
        """The cap must abandon the stream, not buffer the whole body first."""
        chunks_yielded = 0

        class CountingResponse(FakeResponse):
            def iter_content(self, chunk_size: int = 65536):
                nonlocal chunks_yielded
                for _ in range(1000):
                    chunks_yielded += 1
                    yield b"x" * 10

        response = CountingResponse()
        with patch(
            "crewai_tools.security.safe_requests.safe_get", return_value=response
        ):
            with pytest.raises(ValueError):
                safe_get_bounded("https://example.com/huge", max_bytes=25)

        assert chunks_yielded == 3

    def test_error_status_raises(self):
        """An error status propagates as an HTTPError."""
        response = FakeResponse(b"nope", status_code=404)
        with patch(
            "crewai_tools.security.safe_requests.safe_get", return_value=response
        ):
            with pytest.raises(requests.HTTPError):
                safe_get_bounded("https://example.com/missing", max_bytes=1024)

        assert response.closed

    def test_closes_redirect_hops(self):
        """Streamed redirect hops hold connections until closed."""
        hop = FakeResponse(b"", status_code=302)
        response = FakeResponse(b"done")
        response.history = [hop]
        with patch(
            "crewai_tools.security.safe_requests.safe_get", return_value=response
        ):
            safe_get_bounded("https://example.com/start", max_bytes=1024)

        assert hop.closed
        assert response.closed

    def test_requests_are_streamed(self):
        """Streaming is what lets an oversized body be abandoned early."""
        response = FakeResponse(b"ok")
        with patch(
            "crewai_tools.security.safe_requests.safe_get", return_value=response
        ) as safe_get:
            safe_get_bounded("https://example.com/f", max_bytes=1024)

        assert safe_get.call_args.kwargs["stream"] is True
