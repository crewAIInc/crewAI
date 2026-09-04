"""Tool for reading the content at an arbitrary URL as text."""

from __future__ import annotations

import csv
from io import BytesIO, StringIO
from itertools import islice
import re
from typing import Any, Final
from urllib.parse import urlparse
import zipfile

from crewai.tools import BaseTool
from pydantic import BaseModel, Field
import requests

from crewai_tools.security.safe_path import format_error_for_display
from crewai_tools.security.safe_requests import safe_get_bounded


_DEFAULT_MAX_BYTES: Final[int] = 5 * 1024 * 1024
_DEFAULT_TIMEOUT: Final[int] = 30

_PDF_TYPE: Final[str] = "application/pdf"
_DOCX_TYPE: Final[str] = (
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
)
_XLSX_TYPE: Final[str] = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
_HTML_TYPES: Final[frozenset[str]] = frozenset({"text/html", "application/xhtml+xml"})
_TEXT_TYPES: Final[frozenset[str]] = frozenset(
    {
        "application/csv",
        "application/javascript",
        "application/json",
        "application/sql",
        "application/x-ndjson",
        "application/x-yaml",
        "application/xml",
        "application/yaml",
    }
)
_TEXT_TYPE_SUFFIXES: Final[tuple[str, ...]] = ("+json", "+xml", "+yaml")

# Servers commonly serve static files as octet-stream, or send no type at all,
# so the extension and then the body itself are consulted when the header
# carries no usable answer.
_UNINFORMATIVE_TYPES: Final[frozenset[str]] = frozenset(
    {"", "application/octet-stream", "binary/octet-stream"}
)
_EXTENSION_TYPES: Final[dict[str, str]] = {
    ".csv": "text/csv",
    ".docx": _DOCX_TYPE,
    ".htm": "text/html",
    ".html": "text/html",
    ".json": "application/json",
    ".md": "text/markdown",
    ".pdf": _PDF_TYPE,
    ".txt": "text/plain",
    ".xlsx": _XLSX_TYPE,
    ".xml": "application/xml",
    ".yaml": "application/yaml",
    ".yml": "application/yaml",
}

_PDF_MAGIC: Final[bytes] = b"%PDF-"
_ZIP_MAGIC: Final[bytes] = b"PK\x03\x04"
_DOCX_ZIP_ENTRY: Final[str] = "word/document.xml"
_XLSX_ZIP_ENTRY: Final[str] = "xl/workbook.xml"
# A workbook is bounded by max_bytes on the wire but not by what it expands
# into. Two separate ceilings: how much is handed to an agent, and how much
# work a hostile sheet can demand before that decision is even reached.
_XLSX_MAX_CELLS: Final[int] = 200_000
_XLSX_MAX_SCANNED_CELLS: Final[int] = 5_000_000
_HTML_PREFIXES: Final[tuple[str, ...]] = ("<!doctype html", "<html")
_HTML_PREFIX_LEN: Final[int] = max(len(prefix) for prefix in _HTML_PREFIXES)

_SPACES_PATTERN: Final[re.Pattern[str]] = re.compile(r"[ \t]+")
_NEWLINE_PATTERN: Final[re.Pattern[str]] = re.compile(r"\s+\n\s+")


def _charset_from_content_type(content_type: str) -> str | None:
    """Return the charset parameter of a Content-Type header, if it has one."""
    for parameter in content_type.split(";")[1:]:
        name, _, value = parameter.partition("=")
        if name.strip().lower() == "charset":
            return value.strip().strip('"') or None
    return None


class URLReadToolSchema(BaseModel):
    """Input for URLReadTool."""

    url: str = Field(
        ...,
        description=(
            "The http:// or https:// URL to read. Addresses that resolve to "
            "private or internal networks are refused."
        ),
    )
    start_line: int | None = Field(
        1, ge=1, description="Line number to start reading from (1-indexed)"
    )
    line_count: int | None = Field(
        None,
        ge=1,
        description="Number of lines to read. If None, reads the entire content",
    )


class URLReadTool(BaseTool):
    """Read the content at an arbitrary URL and return it as text.

    Unlike :class:`~crewai_tools.tools.file_read_tool.file_read_tool.FileReadTool`,
    which is confined to the local filesystem, this tool performs network
    requests to addresses the caller -- often an LLM -- chooses at runtime. It
    is a separate tool for exactly that reason: granting it is granting network
    egress, and that should be a deliberate choice rather than a flag on a
    filesystem tool.

    Responses are decoded to text according to their content type. PDF, DOCX
    and XLSX bodies have their text extracted, HTML is stripped to visible
    text, and text-shaped types (plain text, Markdown, JSON, XML, YAML, CSV)
    are decoded as-is. When a server names no usable type -- ``octet-stream``
    from a presigned link, say -- the URL extension and then the body's own
    leading bytes are consulted before giving up. Any content that none of
    those identify is refused rather than returned as base64, keeping this
    tool's output text-only.

    That last step is what makes this a good fit for cloud storage. Presigned
    S3 and Cloudflare R2 URLs, and Google Drive, OneDrive and SharePoint
    download links, routinely serve a real document as
    ``application/octet-stream`` from a path that is a content hash with no
    extension, so neither the header nor the URL says what the file is. The
    tool reads a URL and does not authenticate, so the link has to already
    grant access -- which is exactly what a presigned or shared link is.

    Security:
        Requests go through :func:`~crewai_tools.security.safe_requests.safe_get_bounded`,
        which resolves each hostname and rejects it when any resolved address is
        private, loopback, link-local, or otherwise reserved -- covering cloud
        metadata endpoints and internal services. Redirects are never followed
        automatically: every hop is revalidated, and credentials are dropped on
        cross-origin hops. Bodies over ``max_bytes`` are abandoned mid-stream.

        The fetch pins TCP to the IP that passed validation unless
        ``CREWAI_TOOLS_ALLOW_UNSAFE_PATHS`` is set without
        ``CREWAI_TOOLS_FORCE_SAFE_PATHS``. The returned text is still
        untrusted remote content flowing into an agent's context -- a
        fetched page can attempt to instruct the agent. Network egress
        policy and prompt-level handling cover that.

    Args:
        max_bytes (int): Largest response body to accept, in decoded bytes.
            Defaults to 5 MiB.
        timeout (float): Per-request timeout in seconds. Defaults to 30.
        headers (Optional[dict[str, str]]): Extra request headers. Developer
            supplied, not chosen by the model.
        encoding (Optional[str]): Force a text encoding instead of honoring the
            charset the server declares.
        **kwargs: Additional keyword arguments passed to BaseTool.

    Example:
        >>> tool = URLReadTool()
        >>> content = tool.run(url="https://example.com/report.pdf")
        >>> head = tool.run(url="https://example.com/data.csv", line_count=20)
        >>> # A presigned link: no extension, served as octet-stream, read anyway.
        >>> sheet = tool.run(
        ...     url="https://bucket.r2.cloudflarestorage.com/a1b2c3?X-Amz-Signature=..."
        ... )
    """

    name: str = "Read content from a URL"
    description: str = (
        "A tool that reads the content at a URL and returns it as text. To use "
        "this tool, provide a 'url' parameter with an http:// or https:// "
        "address. PDF, DOCX, XLSX, HTML, JSON, XML, CSV and plain-text "
        "responses are converted to text; other binary types are rejected. "
        "Well suited to presigned and share links from S3, Cloudflare R2, "
        "Google Drive, OneDrive and SharePoint, which serve real documents "
        "with a generic content type and no file extension; the type is "
        "detected from the response bytes. The link must already grant "
        "access, as a presigned or shared link does. URLs that resolve to "
        "private or internal network addresses are refused, as are responses "
        "over the tool's size limit. Optionally provide 'start_line' and "
        "'line_count' to read only part of the content."
    )
    args_schema: type[BaseModel] = URLReadToolSchema
    max_bytes: int = _DEFAULT_MAX_BYTES
    timeout: float = _DEFAULT_TIMEOUT
    headers: dict[str, str] | None = None
    encoding: str | None = None

    def __init__(
        self,
        max_bytes: int = _DEFAULT_MAX_BYTES,
        timeout: float = _DEFAULT_TIMEOUT,
        headers: dict[str, str] | None = None,
        encoding: str | None = None,
        **kwargs: Any,
    ) -> None:
        """Initialize the URLReadTool.

        Args:
            max_bytes: Largest response body to accept, in decoded bytes.
            timeout: Per-request timeout in seconds.
            headers: Extra request headers.
            encoding: Force a text encoding instead of the server's charset.
            **kwargs: Additional keyword arguments passed to BaseTool.
        """
        super().__init__(**kwargs)
        self.max_bytes = max_bytes
        self.timeout = timeout
        self.headers = headers
        self.encoding = encoding

    def _request_headers(self) -> dict[str, str]:
        """Return the headers to send, with caller headers taking precedence."""
        return {
            "Accept": "*/*",
            "User-Agent": "Mozilla/5.0 (compatible; crewai-tools URLReadTool)",
            **(self.headers or {}),
        }

    @staticmethod
    def _classify(media_type: str) -> str | None:
        """Map a media type onto the extractor that handles it."""
        if media_type == _PDF_TYPE:
            return "pdf"
        if media_type == _DOCX_TYPE:
            return "docx"
        if media_type == _XLSX_TYPE:
            return "xlsx"
        if media_type in _HTML_TYPES:
            return "html"
        if (
            media_type.startswith("text/")
            or media_type in _TEXT_TYPES
            or media_type.endswith(_TEXT_TYPE_SUFFIXES)
        ):
            return "text"
        return None

    @staticmethod
    def _sniff(body: bytes) -> str | None:
        """Identify a supported type from the body's own bytes.

        Presigned object-store links routinely serve every file as
        octet-stream from an extensionless path, which leaves the bytes as
        the only remaining evidence of what was fetched.

        Fails closed: anything this cannot positively identify is refused
        rather than decoded speculatively, which is what keeps binary
        payloads from reaching an agent's context as mojibake.
        """
        # An empty body identifies nothing. Without this it would strict-decode
        # to "" and read as a successful empty text response.
        if not body:
            return None

        if body.startswith(_PDF_MAGIC):
            return "pdf"

        if body.startswith(_ZIP_MAGIC):
            try:
                with zipfile.ZipFile(BytesIO(body)) as archive:
                    # Central directory only -- nothing is decompressed, so a
                    # zip bomb costs nothing here. Naming the part that must be
                    # present is what keeps a .pptx from reaching python-docx
                    # and surfacing as a misleading "failed to read DOCX"
                    # instead of an honest refusal.
                    names = set(archive.namelist())
            except zipfile.BadZipFile:
                return None
            # Exactly one marker, or nothing: a package claiming to be both a
            # Word document and a workbook has not been positively identified.
            is_docx = _DOCX_ZIP_ENTRY in names
            if is_docx == (_XLSX_ZIP_ENTRY in names):
                return None
            return "docx" if is_docx else "xlsx"

        # Strict, whole-body decode: a prefix would split a multi-byte
        # character and reject valid text. The body is already resident and
        # already capped at max_bytes, so there is nothing to save by slicing.
        try:
            text = body.decode("utf-8")
        except UnicodeDecodeError:
            return None
        if "\x00" in text:
            return None

        leading = text.lstrip("\ufeff").lstrip()[:_HTML_PREFIX_LEN].lower()
        return "html" if leading.startswith(_HTML_PREFIXES) else "text"

    def _resolve_kind(self, body: bytes, content_type: str, *urls: str) -> str | None:
        """Decide how to extract text: content type, URL extension, then bytes.

        Each source is consulted only when every earlier one came back with
        nothing, and none may override an answer an earlier one gave. That
        ordering is what makes the byte sniff a pure widening of what the tool
        accepts: it can turn a refusal into a read, never a read into a
        different read.

        Args:
            body: The fetched body, sniffed last when nothing else identifies it.
            content_type: The raw Content-Type header value.
            *urls: URLs to consult for an extension, most authoritative first.
                A ``.pdf`` link that redirects to an extensionless CDN or
                presigned path only carries its type on the requested URL, so
                both ends of the chain are worth checking.

        Returns:
            The extractor name, or None when nothing identifies the content.
        """
        declared = content_type.split(";", 1)[0].strip().lower()
        if declared not in _UNINFORMATIVE_TYPES:
            return self._classify(declared)

        for url in urls:
            path = urlparse(url).path.lower()
            for extension, media_type in _EXTENSION_TYPES.items():
                if path.endswith(extension):
                    return self._classify(media_type)

        return self._sniff(body)

    def _decode(self, body: bytes, content_type: str) -> str:
        """Decode *body* using the configured, declared, or default encoding.

        Falls back to a replacing UTF-8 decode rather than failing: partially
        readable text is more useful to an agent than an error.
        """
        encoding = self.encoding or _charset_from_content_type(content_type) or "utf-8"
        try:
            return body.decode(encoding)
        except (LookupError, UnicodeDecodeError):
            return body.decode("utf-8", errors="replace")

    @staticmethod
    def _extract_pdf(body: bytes) -> str:
        """Extract text from PDF bytes, page by page."""
        try:
            import pymupdf  # type: ignore[import-untyped]
        except ImportError as e:
            raise ImportError(
                "Reading PDF URLs requires pymupdf. Install with: uv add pymupdf"
            ) from e

        # Opened from memory: the bytes are already in hand, and a temp file
        # would need cleaning up on every error path.
        document = pymupdf.open(stream=body, filetype="pdf")
        try:
            pages = [
                f"Page {number}:\n{text}"
                for number, page in enumerate(document, 1)
                if (text := page.get_text().strip())
            ]
        finally:
            document.close()

        if not pages:
            return "[PDF with no extractable text]"
        return "\n\n".join(pages)

    @staticmethod
    def _extract_docx(body: bytes) -> str:
        """Extract paragraph text from DOCX bytes."""
        try:
            from docx import Document
        except ImportError as e:
            raise ImportError(
                "Reading DOCX URLs requires python-docx. Install with: "
                "uv add python-docx"
            ) from e

        document = Document(BytesIO(body))
        return "\n".join(
            paragraph.text
            for paragraph in document.paragraphs
            if paragraph.text.strip()
        )

    @staticmethod
    def _extract_xlsx(body: bytes) -> str:
        """Extract cell values from XLSX bytes, sheet by sheet, as CSV."""
        try:
            # openpyxl ships no stubs; same treatment as pymupdf above.
            from openpyxl import load_workbook  # type: ignore[import-untyped]
        except ImportError as e:
            raise ImportError(
                "Reading XLSX URLs requires openpyxl. Install with: uv add openpyxl"
            ) from e

        # read_only streams the sheets rather than building the whole object
        # graph, and data_only takes cached values so formulas do not come
        # back as "=SUM(A1:A9)". Both matter for a workbook off an untrusted URL.
        workbook = load_workbook(BytesIO(body), read_only=True, data_only=True)
        try:
            sheets = []
            scannable = _XLSX_MAX_SCANNED_CELLS
            emittable = _XLSX_MAX_CELLS
            truncated = False
            for worksheet in workbook.worksheets:
                # csv rather than a join: a cell may itself hold a comma, a
                # quote or a newline, any of which would corrupt the grid.
                buffer = StringIO()
                writer = csv.writer(buffer, lineterminator="\n")
                for row in worksheet.iter_rows(values_only=True):
                    # Charged before the row is touched. openpyxl pads every
                    # row out to the sheet's declared width, and a forged
                    # dimension makes each *skipped* blank row cost 16k
                    # normalizations -- 4.8 KB of upload drove 1.6e9 of them.
                    # Budgeting only what is emitted bounds none of that work.
                    if len(row) > scannable:
                        truncated = True
                        break
                    scannable -= len(row)

                    values = ["" if value is None else str(value) for value in row]
                    # Excel reports a sheet's dimension generously and openpyxl
                    # pads every row up to it, so one stray far-down cell turns
                    # a 5 KB upload into 100k blank rows. Padding arrives as
                    # None, so testing for exactly-empty drops it while leaving
                    # a cell the author really did fill with spaces alone.
                    if not any(values):
                        continue
                    if len(values) > emittable:
                        truncated = True
                        break
                    emittable -= len(values)
                    writer.writerow(values)

                if buffer.tell():
                    # Only the line terminator: a bare rstrip() would also eat
                    # a trailing space the final cell legitimately holds.
                    grid = buffer.getvalue().rstrip("\n")
                    sheets.append(f"Sheet {worksheet.title}:\n{grid}")
                if truncated:
                    break
        finally:
            workbook.close()

        if not sheets:
            return "[XLSX with no extractable cells]"
        text = "\n\n".join(sheets)
        if truncated:
            text += "\n\n[Truncated: workbook is too large to read in full]"
        return text

    def _extract_html(self, body: bytes, content_type: str) -> str:
        """Strip HTML bytes down to visible text."""
        try:
            from bs4 import BeautifulSoup
        except ImportError as e:
            raise ImportError(
                "Reading HTML URLs requires beautifulsoup4. Install with: "
                "uv add beautifulsoup4"
            ) from e

        soup = BeautifulSoup(self._decode(body, content_type), "html.parser")
        for element in soup(["script", "style"]):
            element.decompose()

        text = _SPACES_PATTERN.sub(" ", soup.get_text(" "))
        return _NEWLINE_PATTERN.sub("\n", text).strip()

    def _extract(self, body: bytes, kind: str, content_type: str) -> str:
        """Dispatch to the extractor named by *kind*."""
        if kind == "pdf":
            return self._extract_pdf(body)
        if kind == "docx":
            return self._extract_docx(body)
        if kind == "xlsx":
            return self._extract_xlsx(body)
        if kind == "html":
            return self._extract_html(body, content_type)
        return self._decode(body, content_type)

    @staticmethod
    def _window(text: str, start_line: int, line_count: int | None) -> str:
        """Return the requested line window of *text*.

        The whole body has already been fetched by this point, so unlike the
        filesystem equivalent this only trims output -- it saves no transfer.

        The bounds are clamped rather than trusted: the args schema rejects
        anything below 1 before it gets here, but islice raises on a negative
        stop index, and this runs outside the caller's error handling.
        """
        if start_line == 1 and line_count is None:
            return text

        start_index = max(start_line - 1, 0)
        stop_index = None if line_count is None else start_index + max(line_count, 0)
        selected = list(islice(text.splitlines(keepends=True), start_index, stop_index))

        if not selected and start_index > 0:
            return (
                f"Error: Start line {start_line} exceeds the number of lines in "
                f"the content."
            )
        return "".join(selected)

    def _run(
        self,
        url: str,
        start_line: int | None = 1,
        line_count: int | None = None,
    ) -> str:
        """Fetch a URL and return its content, or a window of it, as text."""
        start_line = start_line or 1
        line_count = line_count or None

        try:
            body, content_type, final_url = safe_get_bounded(
                url,
                max_bytes=self.max_bytes,
                timeout=self.timeout,
                headers=self._request_headers(),
            )
        except ValueError as e:
            return f"Error: {e}"
        except requests.RequestException as e:
            return f"Error: Failed to fetch '{url}'. {format_error_for_display(e)}"

        kind = self._resolve_kind(body, content_type, final_url, url)
        if kind is None:
            return (
                f"Error: Unsupported content type "
                f"'{content_type.split(';', 1)[0].strip() or 'unknown'}' at "
                f"'{url}'. This tool reads text, HTML, JSON, XML, CSV, PDF and "
                f"DOCX responses."
            )

        try:
            text = self._extract(body, kind, content_type)
        except ImportError as e:
            return f"Error: {e}"
        except Exception as e:
            return (
                f"Error: Failed to read {kind.upper()} content from '{url}'. "
                f"{format_error_for_display(e)}"
            )

        return self._window(text, start_line, line_count)
